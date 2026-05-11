import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from models import IngredientOutput


def generate_excel(results: list[IngredientOutput]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receta Escalada"

    headers = ["Ingrediente", "Gramos (g)", "Kilogramos (kg)"]
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, item in enumerate(results, start=2):
        ws.cell(row=row, column=1, value=item.name)
        ws.cell(row=row, column=2, value=item.grams)
        ws.cell(row=row, column=3, value=item.kilograms)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
