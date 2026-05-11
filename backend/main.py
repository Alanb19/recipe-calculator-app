import io
from contextlib import asynccontextmanager
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from data_loader import load_recipes, Recipe

_recipes: dict[str, Recipe] = {}
_by_name: dict[str, Recipe] = {}  # lowercase name -> Recipe (O(1) lookup)


@asynccontextmanager
async def lifespan(app: FastAPI):
    global _recipes, _by_name
    _recipes = load_recipes()
    seen: set[str] = set()
    for r in _recipes.values():
        if r.code not in seen:
            seen.add(r.code)
            _by_name[r.name.lower()] = r
    yield


app = FastAPI(title="Calculadora de Recetas Nora", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── API ────────────────────────────────────────────────────────────────────────

@app.get("/api/recipes")
def list_recipes(
    search: str = Query("", max_length=200),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    q = search.strip().lower()
    seen: set[str] = set()
    matches = []
    for r in _recipes.values():
        if r.code in seen:
            continue
        seen.add(r.code)
        if not q or q in r.name.lower() or q in r.code:
            matches.append({"code": r.code, "name": r.name, "weight_per_portion": r.weight_per_portion})
    matches.sort(key=lambda r: r["name"])
    return {"total": len(matches), "recipes": matches[offset: offset + limit]}


def _recipe_detail(r: Recipe) -> dict:
    return {
        "code": r.code,
        "name": r.name,
        "weight_per_portion": r.weight_per_portion,
        "components": [
            {"name": c.name, "quantity": c.quantity, "unit": c.unit,
             "type": c.type, "is_principal": c.is_principal}
            for c in r.components
        ],
    }


@app.get("/api/recipes/by-name/{name:path}")
def get_recipe_by_name(name: str):
    key = name.strip().lower()
    # 1. Exact match
    r = _by_name.get(key)
    if r:
        return _recipe_detail(r)
    # 2. Component says "PE X" but recipe is stored as "X" (new fichas parsed without PE prefix)
    if key.startswith("pe "):
        r = _by_name.get(key[3:])
        if r:
            return _recipe_detail(r)
    # 3. Recipe stored as "PE X" but lookup arrives without prefix
    r = _by_name.get("pe " + key)
    if r:
        return _recipe_detail(r)
    raise HTTPException(404, f"Receta {name!r} no encontrada por nombre")


@app.get("/api/recipes/{code}")
def get_recipe(code: str):
    r = _recipes.get(code)
    if not r:
        raise HTTPException(404, f"Receta {code!r} no encontrada")
    return {
        "code": r.code,
        "name": r.name,
        "weight_per_portion": r.weight_per_portion,
        "components": [
            {
                "name": c.name,
                "quantity": c.quantity,
                "unit": c.unit,
                "type": c.type,
                "is_principal": c.is_principal,
            }
            for c in r.components
        ],
    }


class ScaleRequest(BaseModel):
    portions: int = Field(..., ge=1, le=50000)


@app.post("/api/recipes/{code}/scale")
def scale_recipe(code: str, body: ScaleRequest):
    r = _recipes.get(code)
    if not r:
        raise HTTPException(404, f"Receta {code!r} no encontrada")
    return {
        "code": r.code,
        "name": r.name,
        "weight_per_portion": r.weight_per_portion,
        "portions": body.portions,
        "total_weight_kg": round(r.weight_per_portion * body.portions / 1000, 3),
        "components": [
            {
                "name": c.name,
                "quantity_per_portion": c.quantity,
                "quantity_total": round(c.quantity * body.portions, 3),
                "unit": c.unit,
                "type": c.type,
                "is_principal": c.is_principal,
            }
            for c in r.components
        ],
    }


@app.post("/api/recipes/{code}/export")
def export_recipe(code: str, body: ScaleRequest):
    r = _recipes.get(code)
    if not r:
        raise HTTPException(404, f"Receta {code!r} no encontrada")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = r.name[:31]

    # Header info
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    navy = PatternFill("solid", fgColor="192850")
    white_bold = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E6B")

    ws.merge_cells("A1:G1")
    ws["A1"] = f"RECETA: {r.name}  |  Código: {r.code}  |  {body.portions} porciones  |  Peso ref: {r.weight_per_portion}g/porción"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=11)
    ws["A1"].fill = navy
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    headers = ["Nombre", "Tipo", "Cant. x1", "Unidad", f"Cant. x{body.portions}", "Unidad", "Coste total (€)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = white_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    fill_sub = PatternFill("solid", fgColor="EBF0FA")
    fill_ing = PatternFill("solid", fgColor="FFFFFF")

    for i, c in enumerate(r.components, start=3):
        total = round(c.quantity * body.portions, 3)
        row_data = [
            c.name, "Subreceta" if c.type == "subreceta" else "Ingrediente",
            c.quantity, c.unit, total, c.unit,
        ]
        fill = fill_sub if c.type == "subreceta" else fill_ing
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = fill
            cell.border = border
            if col in (3, 5, 7):
                cell.alignment = Alignment(horizontal="right")

    col_widths = [48, 14, 12, 10, 14, 10]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = r.name.replace("/", "-").replace("\\", "-")[:40]
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_name}_{body.portions}p.xlsx"},
    )


# ── Static frontend ────────────────────────────────────────────────────────────
frontend_path = Path(__file__).parent.parent / "frontend"
if frontend_path.exists():
    app.mount("/", StaticFiles(directory=str(frontend_path), html=True), name="frontend")
